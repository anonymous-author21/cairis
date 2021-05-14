#!/usr/bin/python3
#  Licensed to the Apache Software Foundation (ASF) under one
#  or more contributor license agreements.  See the NOTICE file
#  distributed with this work for additional information
#  regarding copyright ownership.  The ASF licenses this file
#  to you under the Apache License, Version 2.0 (the
#  "License"); you may not use this file except in compliance
#  with the License.  You may obtain a copy of the License at
#
#  http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing,
#  software distributed under the License is distributed on an
#  "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY
#  KIND, either express or implied.  See the License for the
#  specific language governing permissions and limitations
#  under the License.

import argparse
import sys
from openpyxl import load_workbook


def main(args=None):
    parser = argparse.ArgumentParser(
        description='Computer Aided Integration of Requirements and Information Security - Workbook to XML converter')
    parser.add_argument('modelFile', help='CAIRIS model file output')
    parser.add_argument('--xlsx', dest='xlsxFile',
                        help='Workbook to input', default='')
    args = parser.parse_args()

    if (args.xlsxFile == ''):
        raise Exception('Workbook file not specified')

    xmlBuf = '<?xml version="1.0"?>\n<!DOCTYPE cairis_model PUBLIC "-//CAIRIS//DTD MODEL 1.0//EN" "https://cairis.org/dtd/cairis_model.dtd">\n\n<cairis_model>\n'

    wb = load_workbook(filename=args.xlsxFile, data_only=True)
    ugSheet = wb.worksheets[0]
    for row in ugSheet.iter_rows(min_row=7):
        ucName = row[0].value
        ucCode = row[1].value
        ucTag = row[2].value
        ucActor = row[3].value
        ucObj = row[4].value
        ucPre = row[5].value
        ucTask = row[6].value
        ucExc = row[7].value
        ucPost = row[8].value
        
        if (ucName != '' and ucName != None and ucName != 0):
            xmlBuf += '  <usecase name = "' + ucName + '" author = "Amna Altaf" code = "' + ucCode + '"> \n '
            xmlBuf += '  <description> ' + ucObj + ' </description> \n '
            xmlBuf += '  <actor name = "' + ucActor + '" /> \n '
            xmlBuf += '  <usecase_environment name = "Morning Shift" >\n '
            xmlBuf += '  <preconditions> ' + ucPre + ' </preconditions> \n '
            xmlBuf += '  <flow> \n '

            if (ucTask != 'Use case ends.')
                count=0    
                xmlBuf += '  <step number = "'count+=1'" description = "' + \ucTask + '"> \n </step> \n '
                min_row+=1

            xmlBuf += '</flow> \n  <postconditions> ' + ucPost + \
                ' </postconditions> \n </usecase_environment> \n </usecase>'

    xmlBuf += "\n</cairis_model>"

    f = open(args.modelFile, 'w')
    f.write(xmlBuf)
    f.close()


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print('Fatal wb2ug error: ' + str(e))
        sys.exit(-1)
